import tkinter as tk
import tkinter.font as tkfont
from tkinter import ttk, filedialog, messagebox, scrolledtext
import json
import os
import re
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.header import Header
from email import encoders
from datetime import datetime
import shutil
import threading
import sys
import subprocess

# ---------------------------------------------------------------------------
# 全局路径常量
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
HISTORY_FILE = os.path.join(BASE_DIR, "history.json")
RESUMES_DIR = os.path.join(BASE_DIR, "resumes")

# ---------------------------------------------------------------------------
# 各邮箱服务商的SMTP服务器地址与端口
# ---------------------------------------------------------------------------
SMTP_SERVERS = {
    "163.com": ("smtp.163.com", 465),
    "126.com": ("smtp.126.com", 465),
    "yeah.net": ("smtp.yeah.net", 465),
    "qq.com": ("smtp.qq.com", 465),
    "foxmail.com": ("smtp.qq.com", 465),
    "gmail.com": ("smtp.gmail.com", 465),
    "outlook.com": ("smtp-mail.outlook.com", 587),
    "hotmail.com": ("smtp-mail.outlook.com", 587),
    "sina.com": ("smtp.sina.com", 465),
    "sohu.com": ("smtp.sohu.com", 465),
}


# ---------------------------------------------------------------------------
# 鼠标悬停提示小部件 - 在目标组件右侧弹出虚线边框白底提示框
# ---------------------------------------------------------------------------
class ToolTip:
    """点击组件时，在其右侧弹出带虚线边框的白色提示框，再次点击任意处关闭。"""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Button-1>", self._toggle)

    def _toggle(self, event=None):
        if self.tip_window:
            self._hide()
        else:
            self._show()

    def _show(self):
        if self.tip_window:
            return
        x = self.widget.winfo_rootx() + self.widget.winfo_width() + 8
        y = self.widget.winfo_rooty()

        tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tw.attributes("-topmost", True)

        canvas = tk.Canvas(tw, bg="#ffffff", highlightthickness=0, bd=0)
        canvas.pack()

        label = tk.Label(
            canvas, text=self.text, justify=tk.LEFT,
            background="#ffffff", font=("{Microsoft YaHei}", 9),
            wraplength=340,
        )
        win_id = canvas.create_window(9, 7, anchor=tk.NW, window=label)
        tw.update_idletasks()
        bbox = canvas.bbox(win_id)
        if bbox:
            x1, y1, x2, y2 = bbox
            canvas.config(width=x2 + 10, height=y2 + 8)
            canvas.create_rectangle(
                x1 - 4, y1 - 3, x2 + 5, y2 + 4,
                outline="#888888", width=1, dash=(4, 3),
            )
        tw.bind("<Button-1>", self._on_tip_click)
        self.tip_window = tw

    def _on_tip_click(self, event):
        self._hide()

    def _hide(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
            self.tip_window = None


# ---------------------------------------------------------------------------
# 工具函数：根据邮箱域名查SMTP配置
# ---------------------------------------------------------------------------
def get_smtp_config(email_addr):
    domain = email_addr.split("@")[-1].lower()
    return SMTP_SERVERS.get(domain, ("smtp." + domain, 465))


# ---------------------------------------------------------------------------
# JSON文件读写工具
# ---------------------------------------------------------------------------
def load_json(path, default=None):
    if default is None:
        default = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------------------
# 配置/历史记录的存取
# ---------------------------------------------------------------------------
def load_config():
    return load_json(CONFIG_FILE, {"email": "", "auth_code": "", "resume_path": ""})


def save_config(config):
    save_json(CONFIG_FILE, config)


def load_history():
    return load_json(HISTORY_FILE, [])


def save_history(records):
    save_json(HISTORY_FILE, records)


# ---------------------------------------------------------------------------
# 公司名称数据库 - 按类别组织，持续扩充
# ---------------------------------------------------------------------------
KNOWN_COMPANIES = [
    # ===== 互联网/科技大厂 =====
    "字节跳动", "今日头条", "抖音", "抖音集团", "TikTok", "TIKTOK", "飞书", "朝夕光年",
    "阿里巴巴", "阿里", "淘宝", "天猫", "支付宝", "蚂蚁集团", "蚂蚁金服", "菜鸟", "阿里云", "高德", "饿了么",
    "腾讯", "微信", "QQ", "腾讯音乐", "腾讯视频", "腾讯游戏",
    "百度", "百度文库", "百度网盘",
    "美团", "美团外卖", "美团优选",
    "京东", "京东物流", "京东健康",
    "拼多多",
    "网易", "网易游戏", "网易云音乐", "有道",
    "小红书", "哔哩哔哩", "B站", "Bilibili",
    "快手", "AcFun", "A站",
    "滴滴", "滴滴出行",
    "小米", "小米汽车", "Redmi",
    "华为", "荣耀", "海思",
    "OPPO", "vivo", "一加", "realme",
    "联想", "摩托罗拉",
    "360", "奇虎360", "奇安信",
    "携程", "去哪儿", "同程旅行",
    "唯品会",
    "Boss直聘", "Boss", "BOSS直聘",
    "知乎",
    "商汤", "商汤科技", "旷视", "旷视科技", "依图", "云从",
    "科大讯飞",
    "大疆", "DJI",
    "深信服",
    "启明星辰",
    "用友", "金蝶",
    "浪潮",
    "紫光", "新华三",
    "中兴", "中兴通讯",
    "传音",
    "Insta360", "影石",
    "万兴科技",
    "光启技术",

    # ===== 金融/证券/银行/保险 =====
    "中金", "中金公司", "CICC",
    "中信证券", "中信建投", "中信建投证券",
    "华泰证券", "华泰联合",
    "国泰君安", "国泰君安证券",
    "海通证券", "海通",
    "广发证券",
    "招商证券",
    "申万宏源", "申万", "宏源",
    "银河证券",
    "东方证券",
    "兴业证券",
    "安信证券",
    "国信证券",
    "光大证券",
    "方正证券",
    "长江证券",
    "平安证券",
    "天风证券",
    "民生证券",
    "国金证券",
    "浙商证券",
    "东吴证券",
    "华创证券",
    "西部证券",
    "东北证券",
    "西南证券",
    "国海证券",
    "中泰证券",
    "中银国际", "中银证券",
    "东兴证券",
    "华西证券",
    "红塔证券",
    "太平洋证券",
    "恒泰证券",
    "开源证券",
    "五矿证券",
    "野村证券", "野村",
    "摩根大通", "JPMorgan", "J.P. Morgan",
    "摩根士丹利", "Morgan Stanley",
    "高盛", "Goldman Sachs", "Goldman",
    "花旗", "Citi", "Citigroup",
    "瑞银", "UBS",
    "瑞信", "Credit Suisse",
    "德意志银行", "Deutsche Bank",
    "汇丰", "HSBC",
    "巴克莱", "Barclays",
    "美银美林", "Bank of America", "BofA",
    "淡马锡", "Temasek",
    "红杉", "红杉资本", "Sequoia",
    "IDG资本", "IDG",
    "高瓴", "高瓴资本", "Hillhouse",
    "经纬中国", "经纬创投",
    "真格基金", "真格",
    "创新工场",
    "顺为资本", "顺为",
    "源码资本",
    "五源资本", "五源",
    "蓝驰创投", "蓝驰",
    "北极光创投", "北极光",
    "金沙江创投", "金沙江",
    "中信资本",
    "鼎晖投资", "鼎晖",
    "弘毅投资", "弘毅",
    "厚朴投资", "厚朴",
    "春华资本", "春华",
    "云锋基金", "云锋",
    "国开金融", "国开",
    "中投公司", "中投",
    "社保基金",
    "华夏基金", "南方基金", "易方达", "博时基金", "嘉实基金", "富国基金",
    "广发基金", "工银瑞信", "招商基金", "鹏华基金", "华安基金", "天弘基金",
    "银华基金", "汇添富", "景顺长城",

    # ===== 外资科技/互联网 =====
    "Google", "谷歌",
    "Microsoft", "微软", "微软中国",
    "Apple", "苹果",
    "Amazon", "亚马逊",
    "Meta", "Facebook",
    "NVIDIA", "英伟达",
    "Intel", "英特尔",
    "AMD",
    "Qualcomm", "高通",
    "Tesla", "特斯拉",
    "Netflix", "奈飞",
    "Airbnb",
    "Uber", "优步",
    "Salesforce",
    "Oracle", "甲骨文",
    "SAP",
    "IBM",
    "PayPal",
    "Shopee", "虾皮",
    "Lazada",
    "Grab",
    "GoTo", "Gojek",

    # ===== 新能源汽车/制造业 =====
    "比亚迪", "BYD",
    "蔚来", "NIO",
    "小鹏", "小鹏汽车", "XPeng",
    "理想", "理想汽车", "Li Auto",
    "吉利", "吉利汽车",
    "长城", "长城汽车",
    "奇瑞",
    "上汽", "上汽集团",
    "一汽", "一汽集团",
    "东风", "东风汽车",
    "广汽", "广汽集团",
    "长安", "长安汽车",
    "江淮", "江淮汽车",
    "宁德时代", "CATL",
    "海康威视", "Hikvision",
    "美的", "格力", "海尔", "TCL",
    "鸿海", "富士康",

    # ===== 游戏/娱乐 =====
    "米哈游", "miHoYo", "HoYoverse",
    "莉莉丝", "莉莉丝游戏",
    "叠纸", "叠纸游戏",
    "鹰角", "鹰角网络",
    "悠星", "悠星网络",
    "库洛", "库洛游戏",
    "散爆", "散爆网络",
    "腾讯光子", "腾讯天美", "腾讯魔方",
    "网易互娱",
    "完美世界",
    "三七互娱",
    "恺英网络",
    "吉比特",
    "祖龙娱乐",
    "IGG",
    "FunPlus", "趣加",
    "沐瞳科技",

    # ===== 电商/本地生活 =====
    "得物", "毒",
    "SHEIN", "希音",
    "Temu",
    "抖音电商",
    "快手电商",
    "山姆", "沃尔玛", "Sam's Club",
    "永辉", "永辉超市",
    "瑞幸", "瑞幸咖啡",

    # ===== 咨询/四大/律所/审计 =====
    "麦肯锡", "McKinsey",
    "波士顿咨询", "BCG",
    "贝恩", "Bain",
    "埃森哲", "Accenture",
    "普华永道", "PwC",
    "德勤", "Deloitte",
    "安永", "EY", "Ernst & Young",
    "毕马威", "KPMG",
    "奥纬", "Oliver Wyman",
    "科尔尼", "AT Kearney", "Kearney",
    "罗兰贝格", "Roland Berger",
    "致同", "Grant Thornton",
    "天职国际",
    "立信",
    "天健",
    "大华",
    "信永中和",
    "中汇",
    "中审众环",

    # ===== 医药/生物 =====
    "恒瑞", "恒瑞医药",
    "药明康德", "药明生物", "WuXi",
    "百济神州", "BeiGene",
    "信达生物",
    "君实生物",
    "康希诺",
    "华大基因", "华大",
    "迈瑞", "迈瑞医疗",
    "联影", "联影医疗",
    "微创", "微创医疗",
    "强生", "Johnson & Johnson",
    "辉瑞", "Pfizer",
    "罗氏", "Roche",
    "诺华", "Novartis",
    "默沙东", "Merck", "MSD",
    "阿斯利康", "AstraZeneca",
    "拜耳", "Bayer",
    "赛诺菲", "Sanofi",
    "葛兰素史克", "GSK",
    "艾伯维", "AbbVie",

    # ===== 地产/建筑 =====
    "万科",
    "碧桂园",
    "恒大",
    "保利", "保利发展",
    "华润", "华润置地",
    "中海", "中海地产",
    "招商蛇口",
    "龙湖", "龙湖集团",
    "绿地",
    "融创",

    # ===== 央企/国企 =====
    "中国石油", "中石油", "PetroChina",
    "中国石化", "中石化", "Sinopec",
    "中国海油", "中海油", "CNOOC",
    "国家电网", "国网",
    "南方电网", "南网",
    "中国移动", "中移动",
    "中国联通", "中联通",
    "中国电信", "中电信",
    "中国建筑", "中建",
    "中国中铁", "中铁",
    "中国铁建", "中铁建",
    "中国交建", "中交",
    "中粮", "中粮集团",
    "中国烟草", "烟草总局",
    "中核集团", "中核",
    "中国航天", "航天科技", "航天科工",
    "中国电子", "CEC",
    "中国电科", "CETC",
    "中国兵器", "兵工集团",
    "中广核",

    # ===== 快递/物流 =====
    "顺丰", "顺丰速运",
    "中通",
    "圆通",
    "韵达",
    "极兔", "J&T",
    "菜鸟网络",

    # ===== 教育/其他 =====
    "新东方",
    "好未来", "学而思",
    "猿辅导",
    "作业帮",
    "高途", "跟谁学",
    "中公教育",
    "华图教育",

    # ===== 快消/零售 =====
    "宝洁", "P&G",
    "联合利华", "Unilever",
    "欧莱雅", "L'Oreal",
    "可口可乐", "Coca-Cola",
    "百事", "Pepsi",
    "雀巢", "Nestle",
    "玛氏", "Mars",
    "耐克", "Nike",
    "阿迪达斯", "Adidas",
    "安踏", "ANTA",
    "李宁", "Li-Ning",

    # ===== 芯片/半导体 =====
    "中芯国际", "SMIC",
    "长江存储", "YMTC",
    "长鑫存储",
    "兆易创新",
    "韦尔股份", "豪威",
    "寒武纪",
    "地平线", "地平线机器人",
    "黑芝麻智能",
    "壁仞科技", "壁仞",
    "摩尔线程",
    "瀚博半导体", "瀚博",
    "沐曦", "沐曦集成电路",
    "天数智芯",
    "景嘉微",

    # ===== AI/大模型 =====
    "DeepSeek", "深度求索",
    "智谱", "智谱AI", "智谱华章",
    "月之暗面", "Moonshot AI",
    "百川智能", "百川",
    "MiniMax", "稀宇科技",
    "零一万物",
    "阶跃星辰",
    "面壁智能", "面壁",
    "无问芯穹",
    "秘塔", "秘塔科技",
]

# 公司后缀模式：能匹配 "XX证券"、"XX银行" 等
COMPANY_SUFFIX_PATTERNS = [
    r"[\w一-鿿一-鿿]{2,8}(?:证券|证券股份有限公司|证券有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:银行|银行股份有限公司|银行有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:基金|基金管理有限公司|基金有限公司|基金公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:保险|保险集团|保险公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:信托|信托有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:期货|期货有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:资本|资本管理有限公司|投资有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:集团|集团有限公司|集团股份有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:私募|私募基金管理有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:科技|科技有限公司|科技股份有限公司)",
    r"[\w一-鿿一-鿿]{2,8}(?:控股|控股集团|控股有限公司)",
]

# 常见城市/省份名（括号匹配时过滤用，避免把【上海】【北京】等地点错当公司）
LOCATION_NAMES = {
    "上海", "北京", "深圳", "广州", "杭州", "成都", "南京", "武汉", "重庆",
    "苏州", "西安", "长沙", "天津", "郑州", "东莞", "青岛", "合肥", "佛山",
    "宁波", "昆明", "沈阳", "大连", "福州", "厦门", "无锡", "济南", "长春",
    "哈尔滨", "石家庄", "南昌", "南宁", "贵阳", "太原", "兰州", "海口", "银川",
    "西宁", "拉萨", "呼和浩特", "乌鲁木齐",
    "北京市", "上海市", "广州市", "深圳市",
    "河北", "山西", "辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽", "福建",
    "江西", "山东", "河南", "湖北", "湖南", "广东", "海南", "四川", "贵州",
    "云南", "陕西", "甘肃", "青海", "台湾", "香港", "澳门",
    "朝阳区", "海淀区", "浦东新区", "徐汇区", "天河区", "南山区",
    "北京朝阳", "北京海淀", "上海浦东", "上海徐汇",
    "全国", "多地", "不限", "远程", "线上", "线下", "总部",
    "中国", "China",
}

# ---------------------------------------------------------------------------
# 智能公司名称提取 - 扫描全文本，多策略匹配
# ---------------------------------------------------------------------------
def extract_company(text):
    """
    从招聘文本中提取公司名称。
    策略（按优先级）：
    1. 已知公司全量扫描：基于公司名称数据库在全文中匹配（最高置信度）
    2. 括号匹配：提取 【】 [] 《》 中的内容（过滤地点名）
    3. 后缀模式匹配：识别 "XX证券" "XX银行" "XX基金" 等
    4. 兜底：从第一行提取
    """
    # --- 策略1：已知公司名称库全文匹配（最高优先级，精确度最高） ---
    sorted_companies = sorted(KNOWN_COMPANIES, key=len, reverse=True)
    for company in sorted_companies:
        if company in text:
            return company

    # --- 策略2：括号提取（过滤地点、非公司内容） ---
    bracket_patterns = [
        r"【\s*(.+?)\s*】",
        r"\[\s*(.+?)\s*\]",
        r"《\s*(.+?)\s*》",
    ]
    for pattern in bracket_patterns:
        m = re.search(pattern, text)
        if m:
            raw = m.group(1).strip()
            if 2 <= len(raw) <= 30 and "@" not in raw:
                # 过滤掉纯地点名
                if raw in LOCATION_NAMES:
                    continue
                # 过滤掉看起来像地点的（如 "上海*1"）
                if any(loc in raw for loc in ["上海", "北京", "深圳", "广州", "杭州"]):
                    if len(raw) <= 5:
                        continue
                return raw

    # --- 策略3：后缀模式匹配 ---
    for pattern in COMPANY_SUFFIX_PATTERNS:
        m = re.search(pattern, text)
        if m:
            return m.group(0)

    # --- 策略4：从第一行兜底提取 ---
    first_line = text.strip().split("\n")[0].strip()
    company = re.sub(r"^(?:招聘|实习|校招|社招|急招|诚聘|诚招|【.*?】|\[.*?\]|\d+\.?)", "", first_line).strip()
    company = re.split(r"[\s\-—|,，、]", company)[0].strip()
    if not company or len(company) < 2:
        parts = re.split(r"\s*[-—|]\s*", first_line, maxsplit=1)
        if len(parts[0]) >= 2 and len(parts[0]) <= 30 and "@" not in parts[0]:
            company = parts[0].strip()

    if 2 <= len(company) <= 30 and "@" not in company:
        return company
    return ""


# ---------------------------------------------------------------------------
# 智能信息提取：从招聘文本中解析邮箱、命名格式、公司、岗位
# 支持多种招聘信息格式
# ---------------------------------------------------------------------------
def extract_info(text):
    result = {"email": "", "naming_format": "", "email_subject": "", "company": "", "position": ""}

    # 提取邮箱
    email_pattern = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"
    emails = re.findall(email_pattern, text)
    if emails:
        result["email"] = emails[0]

    # 提取命名格式 —— 多策略语义理解
    #
    # 占位符关键词：命名格式中表示此处应替换为真实信息的字段名
    PLACEHOLDER_KEYWORDS = [
        "姓名", "名字", "学校", "院校", "大学", "学院", "专业", "方向",
        "岗位", "职位", "应聘岗位", "投递岗位", "实习岗位", "应聘职位",
        "年级", "联系方式", "电话", "手机", "手机号", "邮箱", "邮件",
        "每周", "出勤", "实习时长", "实习周期", "实习时间", "实习期限",
        "到岗时间", "入职时间", "毕业时间", "毕业年份",
        "简历", "个人简历", "应聘", "申请",
    ]
    SEPARATOR_PAT = r"[-_—+/|]"

    def _is_naming_format_like(s):
        """判断字符串是否像命名格式：含分隔符且含占位符关键词。"""
        if not s or len(s) < 3 or len(s) > 80:
            return False
        if not re.search(SEPARATOR_PAT, s):
            return False
        for kw in PLACEHOLDER_KEYWORDS:
            if kw in s:
                return True
        return False

    def _clean_format(s):
        """清理格式字符串：去掉前后标点、残留的引导词前缀。"""
        s = s.strip("\"'\"\"“”「」『』‘’《》（）()[]【】，，。；；：：、、／/ \t")
        # 去掉常见引导词前缀（这些是招聘文案用语，不是命名格式的一部分）
        s = re.sub(
            r"^(?:简历|邮件|文件|附件)?\s*"
            r"(?:(?:及|和|与)?\s*(?:简历|邮件|文件|附件)?\s*"
            r"(?:标题|主题|命名|名称|文件名|格式))?\s*"
            r"(?:请|请以|请用|请将|请按|请按照|请注明|请写明|请标注|"
            r"请统一|请务必|请设置为|请写为|请命名为|请命名为：?|"
            r"请统一为|请以.*?格式|格式.*?为)?\s*"
            r"(?:命名|文件名|名称|格式|标题|主题|格式为|命名为|"
            r"统一格式|命名格式|命名方式|邮件主题|邮件标题|"
            r"简历命名|文件命名|附件命名)?\s*"
            r"[：:是为]+\s*",
            "", s, flags=re.IGNORECASE
        )
        # 去掉尾部句号后的附加说明
        s = re.split(r"[；;。]\s*(?:实习|工作|公司|邮箱|联系|要求|岗位|职责|地址|地点|联系方式)", s)[0]
        # 去掉末尾括号内的说明/备注（示例、命名要求、注意事项等）
        s = re.sub(r'\s*[（(]\s*(?:请严格按|请按照|请按|请|务必|不按|请勿|注意|备注|如|例如|举例|示例|e\.g\.).*$', '', s)
        s = s.rstrip("。；;,.,、 ")
        return s.strip()

    # ---- 策略1：匹配含「分隔符 + 占位符」的格式行（最可靠） ----
    for line in text.split("\n"):
        line = line.strip()
        if not line:
            continue
        # 跳过明显是职责/要求描述的行
        if len(line) > 100:
            continue
        if re.search(r"(职责|描述|要求|负责|协助|参与|完成|具备|熟悉|熟练"
                     r"|开展|撰写|配合|跟进|维护|整理|对接|筹备|支持)",
                     line):
            continue
        # 在行中搜索 "关键词前缀 + 格式内容" 的模式
        kw_prefix = (
            r"(?:简历\s*(?:及|和)?\s*邮件\s*(?:标题|主题|命名|名称|文件名)"
            r"|邮件\s*(?:标题|主题)"
            r"|简历\s*(?:命名|名称|文件名)"
            r"|(?:命名|文件名|名称|标题|主题)\s*(?:格式|方式|要求)?"
            r"|(?:请|请以|请用|请将|请按|请按照|请注明|请写明|请标注|请统一|请务必)"
            r"\s*(?:(?:将|以|按|按照)\s*)?"
            r"(?:简历|邮件|文件名|标题|主题|命名|格式|名称)?"
            r")[：:是为]+\s*"
        )
        m = re.search(kw_prefix + r"(.+)", line, re.IGNORECASE)
        if m:
            raw = _clean_format(m.group(m.lastindex or 1))
            if _is_naming_format_like(raw):
                result["naming_format"] = raw
                break

    # ---- 策略2：匹配引号内的命名格式 ----
    if not result["naming_format"]:
        quote_kw = (
            r"(?:简历|邮件|文件|附件)?\s*"
            r"(?:(?:及|和|与)?\s*(?:简历|邮件|文件|附件)?\s*"
            r"(?:标题|主题|命名|名称|文件名|格式))?\s*"
            r"(?:请|请以|请用|请将|请按|请按照|请注明|请写明|请标注|"
            r"请统一|请务必|请设置为|请写为|请命名为)?\s*"
            r"(?:命名|文件名|名称|格式|标题|主题|格式为|命名为|"
            r"统一格式|命名格式|邮件主题|邮件标题|简历命名|文件命名)?\s*"
            r"[：:是为]?\s*"
        )
        quote_pat = quote_kw + r"[\"“”「」『』‘’《》](.+?)[\"“”「」『』‘’《》]"
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            m = re.search(quote_pat, line, re.IGNORECASE)
            if m:
                fmt = _clean_format(m.group(m.lastindex or 1))
                if 2 < len(fmt) < 60:
                    result["naming_format"] = fmt
                    break

    # ---- 策略3：前策略2的宽松版，匹配关键词后的内容并验证 ----
    if not result["naming_format"]:
        kw_pattern = (
            r"(?:简历|邮件|文件|附件)?\s*"
            r"(?:(?:及|和|与)?\s*(?:简历|邮件|文件|附件)?\s*"
            r"(?:标题|主题|命名|名称|文件名|格式))?\s*"
            r"(?:请|请以|请用|请将|请按|请按照|请注明|请写明|请标注|"
            r"请统一|请务必|请设置为|请写为|请命名为)?\s*"
            r"(?:命名|文件名|名称|格式|标题|主题|格式为|命名为|"
            r"统一格式|命名格式|邮件主题|邮件标题|简历命名|文件命名)"
            r"[：:是为]+\s*"
            r"(.+)"
        )
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            m = re.search(kw_pattern, line, re.IGNORECASE)
            if m:
                fmt = m.group(m.lastindex or 1).strip()
                fmt = _clean_format(fmt)
                if 2 < len(fmt) < 60 and any(
                    c in fmt for c in ["姓", "名", "_", "-", "+", ".", "pdf", "doc",
                                        "简历", "岗位", "公司", "学校", "联系", "电话", "手机"]
                ):
                    result["naming_format"] = fmt
                    break

    # ---- 策略4：兜底 —— 全文中找含分隔符+占位符的行 ----
    if not result["naming_format"]:
        for line in text.split("\n"):
            line = line.strip()
            if not line:
                continue
            if len(line) > 100:
                continue
            if re.search(r"(职责|描述|要求|负责|协助|参与|完成|具备|熟悉|熟练"
                         r"|开展|撰写|配合|跟进|维护|整理|对接|筹备|支持)",
                         line):
                continue
            if re.search(SEPARATOR_PAT, line):
                for kw in PLACEHOLDER_KEYWORDS:
                    if kw in line:
                        pos = line.find(kw)
                        clean = _clean_format(line[pos:])
                        if 3 < len(clean) < 80:
                            result["naming_format"] = clean
                            break
                if result["naming_format"]:
                    break

    # 最终清理
    if result["naming_format"]:
        fmt = result["naming_format"]
        fmt = fmt.strip("\"“”「」『』‘’《》 ")
        fmt = re.sub(r"\s*(?:命名|格式|名称|标题)$", "", fmt)
        result["naming_format"] = fmt.strip()

    # 提取公司名 - 多策略智能匹配（括号 → 已知公司库 → 后缀模式 → 第一行兜底）
    result["company"] = extract_company(text)

    # 提取岗位名 - 优先从标题行提取完整岗位
    lines = text.strip().split("\n")

    # 先在标题行中找完整岗位名
    for line in lines[:3]:
        line = line.strip()
        if not line:
            continue
        cleaned = re.sub(r"【.*?】|\[.*?\]|《.*?》", "", line).strip()
        cleaned = re.sub(r"^(?:招聘|急招|校招|社招|招|诚聘|诚招)\s*", "", cleaned)
        # 匹配完整的岗位名
        POSITION_SUFFIX = r"(?:实习生|工程师|专员|经理|助理|设计师|运营|产品经理|分析师|架构师|开发|管培生|管理培训生|顾问)"
        pos_match = re.search(r"([\w一-鿿（）()]{2,15}" + POSITION_SUFFIX + r")", cleaned)
        if not pos_match:
            # 可能格式为 "岗位 - 公司" 或 "公司 - 岗位"，按减号拆开试两边
            m = re.match(r"^(.+?)\s*[-—|]\s*(.+)$", cleaned)
            if m:
                for part in (m.group(1), m.group(2)):
                    pos_match = re.search(r"([\w一-鿿（）()]{2,15}" + POSITION_SUFFIX + r")", part)
                    if pos_match:
                        cleaned = part
                        break
        if pos_match:
            result["position"] = pos_match.group(1)
            break
        # 宽松匹配：如果整行不长且不含特殊字符，可能就是岗位
        POSITION_BLACKLIST = {"北京", "上海", "广州", "深圳", "杭州", "成都", "武汉",
                               "南京", "远程", "线上", "线下", "实习", "全职", "兼职",
                               "本科", "硕士", "博士", "不限", "若干", "详见", "详情"}
        if (3 <= len(cleaned) <= 18 and "@" not in cleaned and "描述" not in cleaned
                and cleaned not in POSITION_BLACKLIST):
            result["position"] = cleaned
            break

    # 标题没找到，用正则全文搜索
    if not result["position"]:
        POSITION_SUFFIX_FULL = r"(?:实习生|工程师|专员|经理|助理|设计师|开发|运营|产品经理|分析师|架构师|管培生|管理培训生|顾问)"
        position_patterns = [
            r"(?:岗位|职位|招聘|招|找)[：:是为]?\s*([\w一-鿿（）()]{2,15}" + POSITION_SUFFIX_FULL + r")",
            r"([\w一-鿿（）()]{2,15}" + POSITION_SUFFIX_FULL + r")\s*(?:招聘|岗位|职位|JD)",
        ]
        for pattern in position_patterns:
            m = re.search(pattern, text)
            if m:
                pos = m.group(1).strip()
                if 2 <= len(pos) <= 20:
                    result["position"] = pos
                    break

    return result


def parse_resume_info(filename):
    """从简历文件名中提取：实习时长、到岗时间、院校、专业等。"""
    result = {
        "internship_duration": "",
        "arrival_time": "",
        "weekly_attendance": "",
        "undergrad_school": "",
        "grad_school": "",
        "undergrad_major": "",
        "grad_major": "",
        "graduation_time": "",
    }
    # 去掉扩展名
    name_no_ext = os.path.splitext(filename)[0]
    # 按常见分隔符拆分
    segments = re.split(r"[-_@·\s+~—|，、：（）\(\)]+", name_no_ext)
    segments = [s.strip() for s in segments if s.strip()]

    schools = []
    majors = []
    name_segment = segments[0] if segments else ""

    for i, seg in enumerate(segments):
        # 实习时长
        if not result["internship_duration"]:
            CN_NUM = r"(?:[半一两二三四五六七八九十]|十一|十二|\d+)"
            m = re.search(r"实习.{0,3}(?:(?:" + CN_NUM + r")\s*(?:个?月|个月|月|星期|周|天|日))", seg)
            if m:
                result["internship_duration"] = seg
                continue
            if re.search(r"实习", seg):
                result["internship_duration"] = seg
                continue

        # 到岗时间
        if not result["arrival_time"]:
            if re.search(r"(随时|尽快|即刻|立即|一周内|两周内|可|可以|预计|预期|最早|近期).{0,4}(入职|到岗|到)", seg):
                result["arrival_time"] = seg
                continue
            if re.search(r"\d+月.*(入职|到岗|到)", seg):
                result["arrival_time"] = seg
                continue

        # 每周出勤
        if not result["weekly_attendance"]:
            CN_DIGIT = r"(?:[一两二三四五六七八九十]|\d+)"
            if re.search(r"每周.{0,3}" + CN_DIGIT + r"\s*天", seg) or re.search(r"出勤.{0,3}" + CN_DIGIT + r"\s*天", seg):
                result["weekly_attendance"] = seg
                continue

        # 毕业时间
        if not result["graduation_time"]:
            if re.search(r"\d{4}\s*年?\s*\d{0,2}\s*月?\s*(毕业|届)", seg):
                result["graduation_time"] = seg
                continue
            if re.search(r"(应届|往届|毕业(?!院校|学校|生|设计|论))", seg):
                result["graduation_time"] = seg
                continue

        # 院校 - 含"大学/学院/学校"关键词，或常见的校名缩写（纯中文2-10字且非姓名非专业）
        if re.search(r"(大学|学院|学校)", seg):
            if len(seg) <= 20:
                schools.append(seg)
                continue
        # 缩写校名：非首段（非姓名）、纯中文2-10字、不含其他类别关键词
        if i > 0 and re.match(r"^[一-鿿]{2,10}$", seg):
            if not re.search(r"(专业|方向|实习|入职|到岗|月|周|天|\d)", seg):
                schools.append(seg)
                continue

        # 专业
        if re.search(r"(专业|方向|工程|科学|技术|管理|经济|法学|文学|理学|艺术|设计|医学|数学|物理|化学|生物|历史|哲学|新闻|传播|社会|政治|教育|心理|统计|会计|财务|金融)", seg):
            if len(seg) <= 20:
                majors.append(seg)
                continue

    # 分配院校和专业（无明确本科/研究生标记时，猜测第一个为本科）
    if schools:
        for s in schools:
            if re.search(r"研究|硕士|博士|研|硕|博|grad|master|phd", s, re.IGNORECASE):
                if not result["grad_school"]:
                    result["grad_school"] = s
            elif re.search(r"本科|学士|大学|本|undergrad|bachelor", s, re.IGNORECASE):
                if not result["undergrad_school"]:
                    result["undergrad_school"] = s
    # 未明确区分时，第一个归本科，第二个归研究生
    unassigned_schools = [s for s in schools if s not in (result["undergrad_school"], result["grad_school"])]
    for s in unassigned_schools:
        if not result["undergrad_school"]:
            result["undergrad_school"] = s
        elif not result["grad_school"]:
            result["grad_school"] = s

    if majors:
        for m in majors:
            if re.search(r"研究|硕士|博士|研|硕|博|grad|master|phd", m, re.IGNORECASE):
                if not result["grad_major"]:
                    result["grad_major"] = m
            elif re.search(r"本科|学士|本|undergrad|bachelor", m, re.IGNORECASE):
                if not result["undergrad_major"]:
                    result["undergrad_major"] = m
    unassigned_majors = [m for m in majors if m not in (result["undergrad_major"], result["grad_major"])]
    for m in unassigned_majors:
        if not result["undergrad_major"]:
            result["undergrad_major"] = m
        elif not result["grad_major"]:
            result["grad_major"] = m

    return result


# ---------------------------------------------------------------------------
# 邮件发送核心：连接SMTP服务器并发送带附件的邮件
# ---------------------------------------------------------------------------
def send_email(sender_email, auth_code, to_email, subject, body, attachment_path):
    smtp_server, smtp_port = get_smtp_config(sender_email)

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = to_email
    msg["Subject"] = Header(subject, "utf-8")
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(attachment_path, "rb") as f:
        filename = os.path.basename(attachment_path)
        part = MIMEApplication(f.read(), _subtype="pdf", name=filename)
        part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", filename))
        msg.attach(part)

    server = None
    try:
        if smtp_port == 465:
            server = smtplib.SMTP_SSL(smtp_server, smtp_port, timeout=30)
        else:
            server = smtplib.SMTP(smtp_server, smtp_port, timeout=30)
            server.starttls()
        server.login(sender_email, auth_code)
        server.sendmail(sender_email, to_email, msg.as_string())
    finally:
        if server:
            server.quit()


# ---------------------------------------------------------------------------
# 历史记录窗口：查看、搜索、删除、导出投递记录
# ---------------------------------------------------------------------------
class HistoryWindow:
    def __init__(self, parent, records, refresh_callback=None):
        self.parent = parent
        self.records = records
        self.refresh_callback = refresh_callback
        self._item_to_idx = {}  # item_id -> index in self.records
        self.win = tk.Toplevel(parent)
        self.win.title("投递历史记录")
        self.win.geometry("920x520")
        self.win.resizable(True, True)
        self.win.configure(bg="#F0F0F0")
        self.f_heading = tkfont.Font(family="Microsoft YaHei", size=14, weight="bold")
        self.f_body = tkfont.Font(family="Microsoft YaHei", size=10)
        self.f_small = tkfont.Font(family="Microsoft YaHei", size=9)
        self.f_mono = tkfont.Font(family="Consolas", size=9)
        self._build_ui()
        self._refresh_table()

    def _build_ui(self):
        main = tk.Frame(self.win, bg="#F0F0F0")
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # Title
        title_frame = tk.Frame(main, bg="#F0F0F0")
        title_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(title_frame, text="投递历史记录", font=self.f_heading,
                 bg="#F0F0F0").pack(side=tk.LEFT)

        # Search
        search_card = tk.LabelFrame(main, text="搜索", font=self.f_body,
                                    bg="#F0F0F0", padx=12, pady=10)
        search_card.pack(fill=tk.X, pady=(0, 8))

        search_row1 = tk.Frame(search_card, bg="#F0F0F0")
        search_row1.pack(fill=tk.X, pady=(0, 6))
        tk.Label(search_row1, text="关键词", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_row1, textvariable=self.search_var, width=24)
        search_entry.pack(side=tk.LEFT)
        tk.Label(search_row1, text="状态", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(12, 8))
        self.status_filter_var = tk.StringVar(value="全部")
        status_combo = ttk.Combobox(search_row1, textvariable=self.status_filter_var,
                                    values=["全部", "进行中", "已Offer", "已终止"],
                                    state="readonly", width=10)
        status_combo.pack(side=tk.LEFT)
        status_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_table())
        tk.Button(search_row1, text="查询", font=self.f_small,
                  command=self._refresh_table).pack(side=tk.LEFT, padx=(8, 0))
        search_entry.bind("<Return>", lambda e: self._refresh_table())

        search_row2 = tk.Frame(search_card, bg="#F0F0F0")
        search_row2.pack(fill=tk.X)
        tk.Label(search_row2, text="日期范围", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        self.date_start_var = tk.StringVar()
        self.date_start_entry = ttk.Entry(search_row2, textvariable=self.date_start_var, width=12)
        self.date_start_entry.pack(side=tk.LEFT)
        self._setup_date_entry(self.date_start_entry, self.date_start_var, "2026-01-01")
        tk.Label(search_row2, text="—", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=4)
        self.date_end_var = tk.StringVar()
        self.date_end_entry = ttk.Entry(search_row2, textvariable=self.date_end_var, width=12)
        self.date_end_entry.pack(side=tk.LEFT)
        self._setup_date_entry(self.date_end_entry, self.date_end_var, "2026-12-31")
        self.date_start_entry.bind("<Return>", lambda e: self._refresh_table())
        self.date_end_entry.bind("<Return>", lambda e: self._refresh_table())

        # Bottom toolbar — pack before tree so buttons stay visible when window is small
        btn_frame = tk.Frame(main, bg="#F0F0F0")
        btn_frame.pack(fill=tk.X)

        tk.Button(btn_frame, text="删除选中", font=self.f_small,
                  command=self._delete_selected).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_frame, text="导出 CSV", font=self.f_small,
                  command=self._export_csv).pack(side=tk.LEFT, padx=(0, 6))
        tk.Button(btn_frame, text="导出 Excel", font=self.f_small,
                  command=self._export_excel).pack(side=tk.LEFT)

        self.selection_label = tk.Label(btn_frame, text="", font=self.f_small,
                                        fg="#0066CC", bg="#F0F0F0")
        self.selection_label.pack(side=tk.RIGHT, padx=(0, 12))
        hint_label = tk.Label(btn_frame, text="双击投递对象可编辑", font=self.f_small,
                              fg="#999999", bg="#F0F0F0")
        hint_label.pack(side=tk.RIGHT, padx=(0, 6))
        hint_label2 = tk.Label(btn_frame, text="双击所用简历可查看", font=self.f_small,
                               fg="#999999", bg="#F0F0F0")
        hint_label2.pack(side=tk.RIGHT, padx=(0, 12))
        self.count_label = tk.Label(btn_frame, text="", font=self.f_small,
                                     fg="#666666", bg="#F0F0F0")
        self.count_label.pack(side=tk.RIGHT)

        # Table — packed last so it takes remaining space after buttons
        tree_frame = tk.Frame(main, bg="#F0F0F0")
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        style = ttk.Style()
        style.configure("Progress.Treeview", rowheight=46)
        columns = ("object", "email", "time", "resume", "progress")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", selectmode="extended",
                                 style="Progress.Treeview")
        self.tree.heading("object", text="投递对象")
        self.tree.heading("email", text="邮箱")
        self.tree.heading("time", text="投递时间")
        self.tree.heading("resume", text="所用简历 ℹ",
                          command=self._show_resume_path)
        self.tree.heading("progress", text="进度")
        self.tree.column("object", width=200, stretch=False)
        self.tree.column("email", width=140, stretch=False)
        self.tree.column("time", width=130, stretch=False)
        self.tree.column("resume", width=120, stretch=False)
        self.tree.column("progress", width=380, stretch=False)

        tree_scroll_y = tk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        tree_scroll_x = tk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        self.tree.bind("<<TreeviewSelect>>", self._on_selection_change)

        # Drag-to-select
        self._drag_start_row = None
        self._drag_selected = set()
        self.tree.bind("<ButtonPress-1>", self._on_drag_start)
        self.tree.bind("<B1-Motion>", self._on_drag_motion)
        self.tree.bind("<ButtonRelease-1>", self._on_drag_end)
        self.tree.bind("<Shift-ButtonPress-1>", lambda e: "break")

        # Progress color tags (row-level, subtle)
        self.tree.tag_configure("pass", foreground="#27AE60")
        self.tree.tag_configure("fail", foreground="#E74C3C")
        self.tree.tag_configure("pending", foreground="#666666")

        # Double-click: col #1 to edit, col #4 to open resume
        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Motion>", self._on_tree_motion)
        self._edit_entry = None

    def _parse_date(self, text):
        """将用户输入的日期文本解析为 date 对象，支持 YYYY-MM-DD / YYYY/MM/DD / YYYYMMDD。"""
        if not text or not text.strip():
            return None
        text = text.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _setup_date_entry(self, entry, var, placeholder):
        """日期掩码输入框：聚焦时保留横杆，数字变下划线 ____-__-__，输入替换下划线。"""
        MASK = "____-__-__"
        state = {"fresh": False}

        def _on_focus_in(e):
            cur = var.get()
            if cur == placeholder:
                var.set(MASK)
                entry.configure(foreground="black")
                state["fresh"] = True
                entry.after_idle(lambda: entry.icursor(0))

        def _on_focus_out(e):
            cur = var.get()
            if "_" in cur:
                var.set(placeholder)
                entry.configure(foreground="gray")
            state["fresh"] = False

        def _fix_cursor():
            cur = var.get()
            if cur == placeholder or "_" not in cur:
                return
            pos = entry.index(tk.INSERT)
            if pos >= len(cur) or cur[pos] == '-':
                while pos < len(cur) and cur[pos] == '-':
                    pos += 1
                if pos >= len(cur):
                    pos = cur.rfind("_")
                entry.icursor(max(pos, 0))

        def _on_key(e):
            cur = var.get()
            if cur == placeholder:
                var.set(MASK)
                entry.configure(foreground="black")
                state["fresh"] = False
                entry.icursor(0)
                cur = MASK
            if "_" not in cur:
                return

            pos = entry.index(tk.INSERT)
            # 刚从 placeholder 进入掩码，强制光标到最左，确保第一个数字填在年份首位
            if state["fresh"]:
                pos = 0
                entry.icursor(0)
                state["fresh"] = False

            ch = e.char

            if e.keysym == "BackSpace":
                if pos > 0:
                    new_pos = pos - 1
                    while new_pos > 0 and cur[new_pos] == '-':
                        new_pos -= 1
                    if cur[new_pos] != '-':
                        var.set(cur[:new_pos] + "_" + cur[new_pos + 1:])
                        entry.icursor(new_pos)
                return "break"

            if e.keysym == "Delete":
                p = pos
                while p < len(cur) and cur[p] == '-':
                    p += 1
                if p < len(cur) and cur[p] != '-':
                    var.set(cur[:p] + "_" + cur[p + 1:])
                    entry.icursor(p)
                return "break"

            if e.keysym == "Home":
                entry.icursor(0)
                return "break"
            if e.keysym == "End":
                entry.icursor(len(cur))
                return "break"

            if e.keysym == "Left":
                new_pos = pos - 1
                while new_pos > 0 and cur[new_pos] == '-':
                    new_pos -= 1
                entry.icursor(max(new_pos, 0))
                return "break"
            if e.keysym == "Right":
                new_pos = pos + 1
                while new_pos < len(cur) and cur[new_pos] == '-':
                    new_pos += 1
                entry.icursor(min(new_pos, len(cur)))
                return "break"

            if ch and ch.isdigit():
                idx = cur.find("_", pos)
                if idx == -1:
                    return "break"
                var.set(cur[:idx] + ch + cur[idx + 1:])
                next_pos = idx + 1
                while next_pos < len(cur) and cur[next_pos] == '-':
                    next_pos += 1
                entry.icursor(min(next_pos, len(cur) - 1) if next_pos < len(cur) else idx + 1)
                return "break"

            if ch:
                return "break"

        if not hasattr(self, "_placeholders"):
            self._placeholders = {}
        self._placeholders[id(var)] = placeholder

        entry.configure(foreground="gray")
        var.set(placeholder)
        entry.bind("<FocusIn>", _on_focus_in, add="+")
        entry.bind("<FocusOut>", _on_focus_out, add="+")
        entry.bind("<Key>", _on_key, add="+")
        entry.bind("<ButtonRelease-1>", lambda e: entry.after(30, _fix_cursor), add="+")

    def _get_date_val(self, var):
        """获取日期值，如果是 placeholder 或含下划线的掩码则视为空。"""
        val = var.get()
        ph = self._placeholders.get(id(var), "")
        if val == ph or "_" in val:
            return ""
        return val

    # Stage definitions for progress bar
    STAGE_KEYS = ["已投递", "简历通过", "一面", "二面", "三面", "已Offer"]
    STAGE_LABELS = ["投递", "简历", "一面", "二面", "三面", "Offer"]
    STAGE_EMOJI = {"pass": "●", "fail": "✗", "pending": "○"}

    def _get_default_stages(self):
        return {k: "pass" if k == "已投递" else "pending" for k in self.STAGE_KEYS}

    def _get_visible_stages(self, stages):
        """截断阶段列表：遇 fail 停在该处，遇 pending 停在前一处，全部 pass 则全显示。
        返回 (visible_keys, visible_labels)。"""
        cutoff = len(self.STAGE_KEYS) - 1
        for i, key in enumerate(self.STAGE_KEYS):
            state = stages.get(key, "pending")
            if state == "fail":
                cutoff = i
                break
            elif state == "pending":
                cutoff = i - 1
                if cutoff < 0:
                    cutoff = 0
                break
        return self.STAGE_KEYS[:cutoff + 1], self.STAGE_LABELS[:cutoff + 1]

    def _build_progress_text(self, stages):
        """Build the progress bar display text from stages dict."""
        parts = []
        for i, key in enumerate(self.STAGE_KEYS):
            state = stages.get(key, "pending")
            emoji = self.STAGE_EMOJI.get(state, "○")
            parts.append(f"{emoji}{self.STAGE_LABELS[i]}")
        return " ".join(parts)

    def _get_row_tag(self, stages):
        """Determine row color tag based on terminal status."""
        if stages.get("已Offer") == "pass":
            return "pass"
        if "fail" in stages.values():
            return "fail"
        return "pending"

    # ------------------------------------------------------------------
    def _refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._item_to_idx.clear()
        query = self.search_var.get().lower()
        status_filter = self.status_filter_var.get()
        date_start = self._parse_date(self._get_date_val(self.date_start_var))
        date_end = self._parse_date(self._get_date_val(self.date_end_var))
        visible_count = 0
        for idx, r in enumerate(self.records):
            obj = f"{r.get('company', '')} - {r.get('position', '')}"
            email = r.get("email", "")
            resume_file = r.get("resume_file", r.get("naming_format", "") + ".pdf")
            stages = r.get("stages")
            if not stages or not isinstance(stages, dict):
                stages = self._get_default_stages()
                r["stages"] = stages

            # Keyword filter
            if query:
                progress_text = self._build_progress_text(stages)
                if query not in obj.lower() and query not in email.lower() and query not in resume_file.lower() and query not in progress_text.lower():
                    continue

            # Date filter
            if date_start or date_end:
                record_date = self._parse_date(r.get("time", "")[:10])
                if record_date is None:
                    continue
                if date_start and record_date < date_start:
                    continue
                if date_end and record_date > date_end:
                    continue

            # Status filter
            if status_filter == "进行中":
                pending = [k for k in self.STAGE_KEYS if stages.get(k) == "pending"]
                if not pending or pending == self.STAGE_KEYS:
                    continue
            elif status_filter == "已Offer":
                if stages.get("已Offer") != "pass":
                    continue
            elif status_filter == "已终止":
                if "fail" not in stages.values():
                    continue

            progress_text = self._build_progress_text(stages)
            row_tag = self._get_row_tag(stages)
            item_id = self.tree.insert("", tk.END,
                             values=(obj, email, r.get("time", ""), resume_file, progress_text),
                             tags=(row_tag,))
            self._item_to_idx[item_id] = idx
            visible_count += 1
        total = len(self.records)
        if date_start or date_end or query or status_filter != "全部":
            self.count_label.config(text=f"显示 {visible_count} / 共 {total} 条记录")
        else:
            self.count_label.config(text=f"共 {total} 条记录")

    def _on_selection_change(self, event=None):
        n = len(self.tree.selection())
        if n > 0:
            self.selection_label.config(text=f"已选 {n} 条")
        else:
            self.selection_label.config(text="")

    def _on_drag_start(self, event):
        """记录鼠标按下的起始行、Ctrl 状态和已有的选中集合。"""
        self._drag_start_row = self.tree.identify_row(event.y)
        self._drag_moved = False
        self._was_drag = False
        # Ctrl 按下时保留已有选中做追加，否则后续拖拽以起始行到当前行为准
        self._drag_ctrl = bool(event.state & 0x0004)
        if self._drag_ctrl:
            self._drag_base = set(self.tree.selection())
        else:
            self._drag_base = set()

    def _on_drag_motion(self, event):
        """鼠标拖拽时，选中范围 = 起始行 ↔ 当前行（可扩展可收缩），靠近边缘时自动滚动。"""
        if self._drag_start_row is None:
            return
        row = self.tree.identify_row(event.y)

        if row and row != self._drag_start_row:
            self._drag_moved = True
            self._was_drag = True
            self._apply_range(row)
        self._auto_scroll(event)

    def _apply_range(self, row):
        """将选中范围设置为 起始行↔当前行（精确覆盖，双向可收缩）。"""
        all_items = self.tree.get_children("")
        try:
            start_idx = all_items.index(self._drag_start_row)
            end_idx = all_items.index(row)
        except ValueError:
            return
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        wanted = set(all_items[start_idx:end_idx + 1]) | self._drag_base
        current = set(self.tree.selection())
        to_add = wanted - current
        to_remove = current - wanted
        if to_add:
            self.tree.selection_add(*to_add)
        if to_remove:
            self.tree.selection_remove(*to_remove)

    def _auto_scroll(self, event):
        """鼠标靠近 Treeview 顶部或底部时自动滚动。"""
        if hasattr(self, "_scroll_after_id") and self._scroll_after_id:
            self.tree.after_cancel(self._scroll_after_id)
            self._scroll_after_id = None

        height = self.tree.winfo_height()
        if height < 20:
            return
        y = event.y

        if y < 20:
            self.tree.yview_scroll(-1, "units")
            self._scroll_after_id = self.tree.after(80, self._auto_scroll_tick)
        elif y > height - 20:
            self.tree.yview_scroll(1, "units")
            self._scroll_after_id = self.tree.after(80, self._auto_scroll_tick)

    def _auto_scroll_tick(self):
        """定时器回调：边滚动边更新选中范围。"""
        y = self.tree.winfo_pointery() - self.tree.winfo_rooty()
        row = self.tree.identify_row(y)
        if row:
            self._apply_range(row)
        height = self.tree.winfo_height()
        if y < 20:
            self.tree.yview_scroll(-1, "units")
            self._scroll_after_id = self.tree.after(80, self._auto_scroll_tick)
        elif y > height - 20:
            self.tree.yview_scroll(1, "units")
            self._scroll_after_id = self.tree.after(80, self._auto_scroll_tick)

    def _on_drag_end(self, event):
        """清理拖拽状态和自动滚动定时器。"""
        # Progress column click (only if no drag occurred)
        if not getattr(self, '_was_drag', False):
            col = self.tree.identify_column(event.x)
            if col == "#5":
                row_id = self.tree.identify_row(event.y)
                if row_id:
                    bbox = self.tree.bbox(row_id, "#5")
                    if bbox:
                        bx, by, bw, bh = bbox
                        text = self.tree.item(row_id, "values")[4]
                        ft = tkfont.Font(font="TkDefaultFont")
                        text_w = ft.measure(text)
                        pad = 5
                        rel_x = event.x - bx - pad
                        if rel_x <= 0:
                            seg_idx = 0
                        elif rel_x >= text_w:
                            seg_idx = len(self.STAGE_KEYS) - 1
                        else:
                            seg_idx = int(rel_x / text_w * len(self.STAGE_KEYS))
                        self._on_progress_click(row_id, seg_idx)

        if hasattr(self, "_scroll_after_id") and self._scroll_after_id:
            self.tree.after_cancel(self._scroll_after_id)
            self._scroll_after_id = None
        self._drag_start_row = None
        self._drag_moved = False
        self._was_drag = False
        self._drag_ctrl = False
        self._drag_base = set()

    def _on_progress_click(self, item_id, seg_idx):
        """Toggle the stage at seg_idx for the given row, save, and refresh."""
        idx = self._item_to_idx.get(item_id)
        if idx is None or idx < 0 or idx >= len(self.records):
            return
        record = self.records[idx]

        stages = record.get("stages")
        if not stages or not isinstance(stages, dict):
            stages = self._get_default_stages()
            record["stages"] = stages

        key = self.STAGE_KEYS[seg_idx]
        current = stages.get(key, "pending")
        if current == "pending":
            stages[key] = "pass"
        elif current == "pass":
            stages[key] = "fail"
        else:
            stages[key] = "pending"

        save_history(self.records)
        self._refresh_table()

    def _show_resume_path(self):
        dlg = tk.Toplevel(self.win)
        dlg.title("简历存放位置")
        dlg.resizable(False, False)
        dlg.configure(bg="#F0F0F0")
        f = tk.Frame(dlg, bg="#F0F0F0", padx=20, pady=16)
        f.pack()
        tk.Label(f, text="简历副本存放在以下文件夹：", font=self.f_body,
                 bg="#F0F0F0").pack(anchor="w")
        entry = ttk.Entry(f, width=60)
        entry.pack(fill=tk.X, pady=(8, 12))
        entry.insert(0, RESUMES_DIR)
        entry.select_range(0, tk.END)
        entry.focus_set()
        tk.Button(f, text="关闭", font=self.f_small,
                  command=dlg.destroy).pack()
        dlg.transient(self.win)
        dlg.grab_set()
        dlg.geometry("+%d+%d" % (self.win.winfo_rootx() + 100, self.win.winfo_rooty() + 150))
        dlg.wait_window()

    def _on_double_click(self, event):
        col = self.tree.identify_column(event.x)
        if col == "#4":  # 所用简历 column → open PDF
            row_id = self.tree.identify_row(event.y)
            if row_id:
                vals = self.tree.item(row_id, "values")
                if vals and len(vals) > 3:
                    filename = vals[3]
                    path = os.path.join(RESUMES_DIR, filename)
                    if os.path.exists(path):
                        os.startfile(path)
                    else:
                        if messagebox.askyesno("文件不存在",
                                               f"简历文件不存在：\n{path}\n\n是否打开简历存放文件夹？"):
                            os.startfile(RESUMES_DIR)
            return
        if col != "#1":  # 投递对象 column
            return
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return

        bbox = self.tree.bbox(row_id, col)
        if not bbox:
            return
        x, y, w, h = bbox

        vals = self.tree.item(row_id, "values")
        if not vals or len(vals) < 5:
            return
        current_value = vals[0]

        # Create entry widget over the cell
        self._edit_entry = ttk.Entry(self.tree)
        self._edit_entry.place(x=x, y=y, width=w, height=h)
        self._edit_entry.insert(0, current_value)
        self._edit_entry.select_range(0, tk.END)
        self._edit_entry.focus_set()

        self._edit_row_id = row_id

        self._edit_entry.bind("<Return>", self._finish_edit)
        self._edit_entry.bind("<FocusOut>", self._finish_edit)

    def _on_tree_motion(self, event):
        col = self.tree.identify_column(event.x)
        region = self.tree.identify_region(event.x, event.y)
        row_id = self.tree.identify_row(event.y)
        if col == "#1":
            self.tree.config(cursor="xterm")
        elif col == "#4" and (row_id or region == "heading"):
            self.tree.config(cursor="hand2")
        else:
            self.tree.config(cursor="")

    def _finish_edit(self, event=None):
        if not self._edit_entry:
            return
        new_value = self._edit_entry.get().strip()
        self._edit_entry.destroy()
        self._edit_entry = None

        row_id = getattr(self, "_edit_row_id", None)
        if not row_id or not new_value:
            return

        idx = self._item_to_idx.get(row_id)
        if idx is None or idx < 0 or idx >= len(self.records):
            return
        record = self.records[idx]

        # Parse "company - position" back to separate fields
        if " - " in new_value:
            company, position = new_value.split(" - ", 1)
        else:
            company, position = new_value, ""

        record["company"] = company.strip()
        record["position"] = position.strip()
        save_history(self.records)
        self._refresh_table()
        if self.refresh_callback:
            self.refresh_callback()

    def _delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            return
        if not messagebox.askyesno("确认", f"确定删除选中的 {len(selected)} 条记录吗？"):
            return
        indices_to_delete = set()
        for sel in selected:
            idx = self._item_to_idx.get(sel)
            if idx is not None:
                indices_to_delete.add(idx)
        self.records = [
            r for i, r in enumerate(self.records) if i not in indices_to_delete
        ]
        save_history(self.records)
        self._refresh_table()
        if self.refresh_callback:
            self.refresh_callback()

    def _export_csv(self):
        path = filedialog.asksaveasfilename(
            title="导出历史记录",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv")],
            initialfile="投递历史记录.csv",
        )
        if not path:
            return
        import csv
        all_stage_keys = self.STAGE_KEYS
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["投递对象", "邮箱", "投递时间", "所用简历"] + all_stage_keys)
            for r in self.records:
                stages = r.get("stages")
                if not stages or not isinstance(stages, dict):
                    stages = self._get_default_stages()
                visible_keys, _ = self._get_visible_stages(stages)
                stage_vals = []
                for k in visible_keys:
                    s = stages.get(k, "pending")
                    stage_vals.append("✓" if s == "pass" else ("✗" if s == "fail" else "-"))
                # 补齐不足的列（空字符串）
                while len(stage_vals) < len(all_stage_keys):
                    stage_vals.append("")
                writer.writerow([
                    f"{r.get('company','')} - {r.get('position','')}",
                    r.get("email", ""),
                    r.get("time", ""),
                    r.get("resume_file", r.get("naming_format", "") + ".pdf"),
                ] + stage_vals)
        messagebox.showinfo("导出成功", f"已导出到:\n{path}")

    def _export_excel(self):
        path = filedialog.asksaveasfilename(
            title="导出历史记录",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile="投递历史记录.xlsx",
        )
        if not path:
            return
        import openpyxl
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        all_stage_keys = self.STAGE_KEYS
        green_font = Font(color="27AE60", bold=True)
        red_font = Font(color="E74C3C", bold=True)
        ws.append(["公司", "岗位", "邮箱", "投递时间", "所用简历"] + all_stage_keys)
        for r in self.records:
            stages = r.get("stages")
            if not stages or not isinstance(stages, dict):
                stages = self._get_default_stages()
            visible_keys, _ = self._get_visible_stages(stages)
            stage_vals = []
            for k in visible_keys:
                s = stages.get(k, "pending")
                stage_vals.append((s, "✓" if s == "pass" else ("✗" if s == "fail" else "-")))
            while len(stage_vals) < len(all_stage_keys):
                stage_vals.append(("", ""))
            row_data = [
                r.get("company", ""),
                r.get("position", ""),
                r.get("email", ""),
                r.get("time", ""),
                r.get("resume_file", r.get("naming_format", "") + ".pdf"),
            ]
            ws.append(row_data + [v for _, v in stage_vals])
            row_idx = ws.max_row
            col_offset = len(row_data) + 1
            for i, (state, _) in enumerate(stage_vals):
                cell = ws.cell(row=row_idx, column=col_offset + i)
                if state == "pass":
                    cell.font = green_font
                elif state == "fail":
                    cell.font = red_font
        wb.save(path)
        messagebox.showinfo("导出成功", f"已导出到:\n{path}")


# ---------------------------------------------------------------------------
# 主应用窗口：整合配置、解析、邮件发送等功能
# ---------------------------------------------------------------------------
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("智能投递助手")
        self.root.geometry("720x680")
        self.root.resizable(True, True)
        self.root.configure(bg="#F0F0F0")
        self.root.option_add("*Font", "{Microsoft YaHei} 10")

        # ---- 字体 ----
        self.f_title = tkfont.Font(family="Microsoft YaHei", size=16, weight="bold")
        self.f_heading = tkfont.Font(family="Microsoft YaHei", size=14, weight="bold")
        self.f_small = tkfont.Font(family="Microsoft YaHei", size=9)
        self.f_body = tkfont.Font(family="Microsoft YaHei", size=10)
        self.f_mono = tkfont.Font(family="Consolas", size=9)

        self.config = load_config()

        # 个人信息默认值
        self.personal = self.config.get("personal", {})
        self.personal.setdefault("name", "")
        self.personal.setdefault("phone", "")
        self.personal.setdefault("school", "")
        self.personal.setdefault("internship_duration", "")
        self.personal.setdefault("arrival_time", "")
        self.personal.setdefault("weekly_attendance", "")
        self.personal.setdefault("undergrad_school", "")
        self.personal.setdefault("grad_school", "")
        self.personal.setdefault("undergrad_major", "")
        self.personal.setdefault("grad_major", "")
        self.personal.setdefault("graduation_time", "")
        self.personal.setdefault("grade", "")

        self.records = load_history()

        # 解析已保存的邮箱，拆分为用户名和域名
        saved_email = self.config.get("email", "")
        if "@" in saved_email:
            saved_user, saved_domain = saved_email.split("@", 1)
            saved_domain = "@" + saved_domain
        else:
            saved_user, saved_domain = "", ""

        self.resume_path = tk.StringVar(value=self.config.get("resume_path", ""))
        self.naming_format = tk.StringVar()
        self.raw_naming_format = ""  # 未填充占位符的原始命名格式，避免 _send 二次填充
        self.to_email = tk.StringVar()
        self.email_user = tk.StringVar(value=saved_user)
        self.email_domain = tk.StringVar(value=saved_domain)
        self.sender_email = tk.StringVar(value=saved_email)
        self.auth_code = tk.StringVar(value=self.config.get("auth_code", ""))
        self.company_var = tk.StringVar()
        self.position_var = tk.StringVar()
        self.email_subject = tk.StringVar(value="实习申请")
        self.user_name = tk.StringVar(value=self.personal.get("name", ""))
        self.user_phone = tk.StringVar(value=self.personal.get("phone", ""))
        self.user_school = tk.StringVar(value=self.personal.get("school", ""))
        self.internship_duration = tk.StringVar(value=self.personal.get("internship_duration", ""))
        self.arrival_time = tk.StringVar(value=self.personal.get("arrival_time", ""))
        self.weekly_attendance = tk.StringVar(value=self.personal.get("weekly_attendance", ""))
        self.undergrad_school = tk.StringVar(value=self.personal.get("undergrad_school", ""))
        self.grad_school = tk.StringVar(value=self.personal.get("grad_school", ""))
        self.undergrad_major = tk.StringVar(value=self.personal.get("undergrad_major", ""))
        self.grad_major = tk.StringVar(value=self.personal.get("grad_major", ""))
        self.graduation_time = tk.StringVar(value=self.personal.get("graduation_time", ""))
        self.user_grade = tk.StringVar(value=self.personal.get("grade", ""))

        self.email_domains = [
            "@163.com", "@126.com", "@yeah.net", "@qq.com", "@foxmail.com",
            "@gmail.com", "@outlook.com", "@hotmail.com", "@sina.com", "@sohu.com",
            "@aliyun.com", "@139.com", "@189.cn",
        ]

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # 首次启动自动弹出使用指南
        if not self.config.get("tutorial_shown"):
            self.root.after(300, self._show_tutorial)
            self.config["tutorial_shown"] = True
            save_config(self.config)

        # 重试状态
        self._retry_state = None   # 当前重试参数
        self._retry_after_id = None  # root.after 的 id

    def _build_ui(self):
        # ---- scrollable canvas wrapper ----
        canvas = tk.Canvas(self.root, bg="#F0F0F0", highlightthickness=0)
        scrollbar = tk.Scrollbar(self.root, orient=tk.VERTICAL, command=canvas.yview)
        self.scroll_frame = tk.Frame(canvas, bg="#F0F0F0")

        self.scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        self._canvas_window = canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")

        def _on_canvas_configure(event):
            canvas.itemconfig(self._canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        def _bind_mw(e): canvas.bind_all("<MouseWheel>", _on_mousewheel)
        def _unbind_mw(e): canvas.unbind_all("<MouseWheel>")
        canvas.bind("<Enter>", _bind_mw)
        canvas.bind("<Leave>", _unbind_mw)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # ====== main container ======
        main = tk.Frame(self.scroll_frame, bg="#F0F0F0")
        main.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

        # ---- Title ----
        title_frame = tk.Frame(main, bg="#F0F0F0")
        title_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(title_frame, text="智能投递助手", font=self.f_title,
                 bg="#F0F0F0").pack(side=tk.LEFT)
        tk.Button(title_frame, text="使用指南", font=self.f_small,
                  command=self._show_tutorial).pack(side=tk.LEFT, padx=(8, 0))
        tk.Button(title_frame, text="历史记录", font=self.f_small,
                  command=self._open_history).pack(side=tk.RIGHT)

        # ================================================================
        # Section 1 — 发件配置 (LabelFrame with visible border)
        # ================================================================
        c1 = tk.LabelFrame(main, text="发件配置", font=self.f_heading,
                           bg="#F0F0F0", padx=12, pady=10)
        c1.pack(fill=tk.X, pady=(0, 8))

        row1 = tk.Frame(c1, bg="#F0F0F0")
        row1.pack(fill=tk.X)
        tk.Label(row1, text="邮箱", font=self.f_body, bg="#F0F0F0").grid(row=0, column=0, sticky=tk.W, padx=(0, 6))
        email_inner = tk.Frame(row1, bg="#F0F0F0")
        email_inner.grid(row=0, column=1, sticky=tk.W)
        self.email_entry = ttk.Entry(email_inner, textvariable=self.email_user, width=16)
        self.email_entry.pack(side=tk.LEFT)
        self.email_entry.bind("<KeyRelease>", lambda e: self._update_sender_email())
        self.email_combo = ttk.Combobox(
            email_inner, textvariable=self.email_domain, values=self.email_domains,
            state="readonly", width=14,
        )
        self.email_combo.pack(side=tk.LEFT, padx=(4, 0))
        self.email_combo.bind("<<ComboboxSelected>>", self._on_domain_selected)
        saved_domain = self.email_domain.get()
        if saved_domain and saved_domain in self.email_domains:
            self.email_combo.set(saved_domain)
        elif saved_domain:
            self.email_combo.set(saved_domain)

        tk.Label(row1, text="授权码", font=self.f_body, bg="#F0F0F0").grid(row=0, column=2, sticky=tk.W, padx=(20, 6))
        auth_frame = tk.Frame(row1, bg="#F0F0F0")
        auth_frame.grid(row=0, column=3, sticky=tk.W)
        ttk.Entry(auth_frame, textvariable=self.auth_code, width=24, show="•").pack(side=tk.LEFT)
        auth_info_lbl = tk.Label(
            auth_frame, text="ⓘ", fg="#0066CC",
            font=("{Microsoft YaHei}", 10, "bold"), cursor="hand2",
            bg="#F0F0F0",
        )
        auth_info_lbl.pack(side=tk.LEFT, padx=(4, 0))
        ToolTip(auth_info_lbl, (
            "授权码是邮箱服务商为第三方客户端（如本软件）提供的\n"
            "专用登录密码，不是您的邮箱登录密码。\n\n"
            "获取方式：\n"
            "● 163/126/yeah邮箱：登录网页邮箱 → 设置 →\n"
            "   POP3/SMTP/IMAP → 开启SMTP服务 → 生成授权码\n"
            "● QQ/foxmail邮箱：登录网页邮箱 → 设置 → 账户 →\n"
            "   POP3/IMAP/SMTP服务 → 生成授权码\n"
            "● Gmail邮箱：Google账户 → 安全性 → 应用专用密码\n"
            "  （需先开启两步验证）\n"
            "● Outlook/Hotmail：Microsoft账户 → 安全性 →\n"
            "   应用密码"
        ))

        # ================================================================
        # Section 2 — 个人信息
        # ================================================================
        c2 = tk.LabelFrame(main, text="个人信息", font=self.f_heading,
                           bg="#F0F0F0", padx=12, pady=10)
        c2.pack(fill=tk.X, pady=(0, 8))
        tk.Label(c2, text="用于自动填充简历命名模板（以下均为选填）",
                 font=self.f_small, fg="#666666", bg="#F0F0F0").pack(anchor=tk.W, pady=(0, 8))

        grid = tk.Frame(c2, bg="#F0F0F0")
        grid.pack(fill=tk.X)
        fields = [
            [("姓名", self.user_name, 12), ("手机号", self.user_phone, 14), ("学校", self.user_school, 12)],
            [("本科院校", self.undergrad_school, 12), ("本科专业", self.undergrad_major, 14), ("研究生院校", self.grad_school, 12)],
            [("研究生专业", self.grad_major, 12), ("实习时长", self.internship_duration, 14), ("到岗时间", self.arrival_time, 12)],
            [("每周出勤", self.weekly_attendance, 12), ("毕业时间", self.graduation_time, 12), ("年级", self.user_grade, 12)],
        ]
        for r, row_fields in enumerate(fields):
            for c_idx, (label, var, w) in enumerate(row_fields):
                col_offset = c_idx * 2
                tk.Label(grid, text=label, font=self.f_body, bg="#F0F0F0").grid(
                    row=r, column=col_offset, sticky=tk.W, padx=(0 if c_idx == 0 else 14, 4), pady=(0 if r == 0 else 8, 0))
                ttk.Entry(grid, textvariable=var, width=w).grid(
                    row=r, column=col_offset + 1, sticky=tk.W, pady=(0 if r == 0 else 8, 0))

        # ================================================================
        # Section 3 — 简历文件
        # ================================================================
        c3 = tk.LabelFrame(main, text="原始简历", font=self.f_heading,
                           bg="#F0F0F0", padx=12, pady=10)
        c3.pack(fill=tk.X, pady=(0, 8))
        resume_row = tk.Frame(c3, bg="#F0F0F0")
        resume_row.pack(fill=tk.X)
        ttk.Entry(resume_row, textvariable=self.resume_path, width=42).pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(resume_row, text="选择 PDF", font=self.f_small,
                  command=self._select_resume).pack(side=tk.LEFT, padx=(8, 0))

        # ================================================================
        # Section 4 — 提取结果
        # ================================================================
        c4 = tk.LabelFrame(main, text="提取结果", font=self.f_heading,
                           bg="#F0F0F0", padx=12, pady=10)
        c4.pack(fill=tk.X, pady=(0, 8))

        f1 = tk.Frame(c4, bg="#F0F0F0")
        f1.pack(fill=tk.X, pady=(0, 6))
        tk.Label(f1, text="邮件标题", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(f1, textvariable=self.email_subject, width=48).pack(side=tk.LEFT, fill=tk.X, expand=True)

        f2 = tk.Frame(c4, bg="#F0F0F0")
        f2.pack(fill=tk.X, pady=(0, 6))
        tk.Label(f2, text="公司", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(f2, textvariable=self.company_var, width=18).pack(side=tk.LEFT)
        tk.Label(f2, text="岗位", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(18, 8))
        ttk.Entry(f2, textvariable=self.position_var, width=18).pack(side=tk.LEFT)

        f3 = tk.Frame(c4, bg="#F0F0F0")
        f3.pack(fill=tk.X, pady=(0, 6))
        tk.Label(f3, text="对方邮箱", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(f3, textvariable=self.to_email, width=48).pack(side=tk.LEFT, fill=tk.X, expand=True)

        f4 = tk.Frame(c4, bg="#F0F0F0")
        f4.pack(fill=tk.X)
        tk.Label(f4, text="命名格式", font=self.f_body, bg="#F0F0F0").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Entry(f4, textvariable=self.naming_format, width=48).pack(side=tk.LEFT, fill=tk.X, expand=True)

        # ================================================================
        # Section 5 — 招聘信息
        # ================================================================
        c5 = tk.LabelFrame(main, text="招聘信息", font=self.f_heading,
                           bg="#F0F0F0", padx=12, pady=10)
        c5.pack(fill=tk.X, pady=(0, 8))
        self.job_text = scrolledtext.ScrolledText(
            c5, height=9, width=70,
            font=self.f_body,
            bg="#FFFFFF", fg="#000000", relief="sunken", borderwidth=1,
        )
        self.job_text.pack(fill=tk.X)

        # ================================================================
        # Action buttons
        # ================================================================
        btn_frame = tk.Frame(main, bg="#F0F0F0")
        btn_frame.pack(fill=tk.X, pady=(2, 8))
        tk.Button(btn_frame, text="智能提取", font=self.f_body, width=10,
                  command=self._extract).pack(side=tk.LEFT, padx=(0, 8))
        self.send_btn = tk.Button(btn_frame, text="一键发送", font=self.f_body, width=10,
                                   command=self._send)
        self.send_btn.pack(side=tk.LEFT)
        self.cancel_retry_btn = tk.Button(btn_frame, text="取消重试", font=self.f_small,
                                           command=self._cancel_retry)

        # ================================================================
        # Log
        # ================================================================
        log_frame = tk.LabelFrame(main, text="日志", font=self.f_heading,
                                  bg="#F0F0F0", padx=12, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_area = scrolledtext.ScrolledText(
            log_frame, height=5, width=70,
            font=self.f_mono,
            bg="#FFFFFF", fg="#000000", relief="sunken", borderwidth=1,
            state=tk.DISABLED,
        )
        self.log_area.pack(fill=tk.BOTH, expand=True)

    def _log(self, msg):
        self.log_area.config(state=tk.NORMAL)
        self.log_area.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_area.see(tk.END)
        self.log_area.config(state=tk.DISABLED)

    def _fill_template(self, fmt, info):
        """将命名格式中的占位符按语义替换为真实信息。

        使用单次正则替换，按从长到短顺序匹配占位符，长词优先，
        已替换部分不会被后续短词再次匹配（避免"复旦大学"中的"大学"被二次替换）。
        """
        school = (
            self.grad_school.get().strip() or
            self.user_school.get().strip()
        )
        undergrad_school = self.undergrad_school.get().strip()
        major = self.grad_major.get().strip()
        undergrad_major = self.undergrad_major.get().strip()
        weekly = self.weekly_attendance.get().strip()
        duration = self.internship_duration.get().strip()
        arrival = self.arrival_time.get().strip()
        grad_time = self.graduation_time.get().strip()
        name = self.user_name.get().strip()
        phone = self.user_phone.get().strip()
        email = self.sender_email.get().strip()
        grade = self.user_grade.get().strip()
        position = info.get("position", "")
        company = info.get("company", "")

        # 全部候选，按长度降序排列（长占位符优先，避免短词截胡）
        # 格式: [(占位符, 替换值), ...]
        candidates = [
            # ===== 院校（研究生优先，其次通用学校）=====
            ("研究生院校", self.grad_school.get().strip()),
            ("硕士院校", self.grad_school.get().strip()),
            ("毕业院校", school),
            ("在读院校", school),
            ("所在院校", school),
            ("就读院校", school),
            ("硕士学校", self.grad_school.get().strip()),
            ("本科院校", undergrad_school),
            ("本科学校", undergrad_school),
            ("本科学院", undergrad_school),
            ("毕业学校", school),
            ("在读学校", school),
            ("所在学校", school),
            ("就读学校", school),
            ("院校", school),
            ("大学", school),
            ("学院", school),
            ("学校", school),
            # ===== 专业（仅硕士专业）=====
            ("研究生专业", self.grad_major.get().strip()),
            ("硕士专业", self.grad_major.get().strip()),
            ("本科专业", undergrad_major),
            ("所学专业", major),
            ("主修专业", major),
            ("专业", major),
            ("方向", major),
            # ===== 岗位 / 职位 =====
            ("应聘岗位", position),
            ("投递岗位", position),
            ("实习岗位", position),
            ("申请岗位", position),
            ("求职岗位", position),
            ("意向岗位", position),
            ("目标岗位", position),
            ("应聘职位", position),
            ("投递职位", position),
            ("申请职位", position),
            ("求职职位", position),
            ("意向职位", position),
            ("期望岗位", position),
            ("期望职位", position),
            ("招聘岗位", position),
            ("岗位", position),
            ("职位", position),
            # ===== 姓名 =====
            ("中文姓名", name),
            ("你的姓名", name),
            ("姓名", name),
            ("名字", name),
            ("中文名", name),
            # ===== 每周出勤（具体到泛化）=====
            ("每周可出勤天数", weekly),
            ("每周可实习天数", weekly),
            ("每周出勤天数", weekly),
            ("每周工作天数", weekly),
            ("每周实习天数", weekly),
            ("每周到岗天数", weekly),
            ("每周上班天数", weekly),
            ("每周工作几天", weekly),
            ("可实习天数", weekly),
            ("每周出勤几天", weekly),
            ("每周到岗几天", weekly),
            ("每周实习几天", weekly),
            ("每周几天出勤", weekly),
            ("每周几天到岗", weekly),
            ("每周几天", weekly),
            ("实习几天", weekly),
            ("每周天数", weekly),
            ("出勤几天", weekly),
            ("几天出勤", weekly),
            ("周出勤天数", weekly),
            ("每周出勤", weekly),
            ("出勤天数", weekly),
            ("出勤时间", weekly),
            ("每周到岗", weekly),
            ("出勤", weekly),
            ("每周", weekly),
            # ===== 实习时长 / 周期 =====
            ("计划实习月数", duration),
            ("可持续实习时间", duration),
            ("可实习时长", duration),
            ("实习周期", duration),
            ("实习期限", duration),
            ("实习时长", duration),
            ("实习时间", duration),
            ("实习月数", duration),
            ("实习期", duration),
            ("可实习时间", duration),
            # ===== 到岗 / 入职 =====
            ("预期到岗时间", arrival),
            ("预计到岗时间", arrival),
            ("最早到岗时间", arrival),
            ("可到岗时间", arrival),
            ("到岗时间", arrival),
            ("到岗日期", arrival),
            ("入职时间", arrival),
            ("入职日期", arrival),
            # ===== 毕业时间 =====
            ("预计毕业时间", grad_time),
            ("预期毕业时间", grad_time),
            ("毕业时间", grad_time),
            ("毕业年份", grad_time),
            ("毕业日期", grad_time),
            # ===== 年级 =====
            ("在读年级", grade),
            ("所在年级", grade),
            ("年级", grade),
            # ===== 联系方式 / 电话 =====
            ("联系电话", phone),
            ("手机号码", phone),
            ("电话号码", phone),
            ("联系方式", phone),
            ("联系手机", phone),
            ("手机号", phone),
            ("电话", phone),
            ("手机", phone),
            # ===== 邮箱 =====
            ("电子邮箱", email),
            ("邮箱地址", email),
            ("联系邮箱", email),
            ("个人邮箱", email),
            ("E-mail", email),
            ("Email", email),
            ("邮箱", email),
            ("邮件", email),
            # ===== 公司 =====
            ("公司名称", company),
            ("公司", company),
        ]

        # 过滤掉占位符为空或替换值为空的候选项
        active = [(p, v) for p, v in candidates if p and v]
        # 按占位符长度降序排列
        active.sort(key=lambda x: -len(x[0]))

        if not active:
            return fmt

        # 构建正则：长词在前，同一位置优先匹配更长的占位符
        pattern = "|".join(re.escape(p) for p, _ in active)
        value_map = {p: v for p, v in active}

        def _replacer(m):
            return value_map.get(m.group(0), m.group(0))

        return re.sub(pattern, _replacer, fmt)

    def _on_domain_selected(self, event=None):
        """域名下拉选择时更新完整邮箱地址。"""
        self._update_sender_email()

    def _update_sender_email(self):
        user = self.email_user.get().strip()
        domain = self.email_domain.get().strip()
        if user and domain:
            self.sender_email.set(f"{user}{domain}")
        elif domain:
            self.sender_email.set(domain)
        else:
            self.sender_email.set(user)

    def _select_resume(self):
        path = filedialog.askopenfilename(
            title="选择简历文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")],
        )
        if path:
            self.resume_path.set(path)
            self.config["resume_path"] = path
            save_config(self.config)
            self._auto_fill_from_resume(path)

    def _auto_fill_from_resume(self, path):
        """选择简历文件后自动从文件名解析信息并回填到个人信息字段。"""
        info = parse_resume_info(os.path.basename(path))
        updated = []
        for field, var, label in [
            ("internship_duration", self.internship_duration, "实习时长"),
            ("arrival_time", self.arrival_time, "到岗时间"),
            ("weekly_attendance", self.weekly_attendance, "每周出勤"),
            ("undergrad_school", self.undergrad_school, "本科院校"),
            ("grad_school", self.grad_school, "研究生院校"),
            ("undergrad_major", self.undergrad_major, "本科专业"),
            ("grad_major", self.grad_major, "研究生专业"),
            ("graduation_time", self.graduation_time, "毕业时间"),
        ]:
            if info[field]:
                var.set(info[field])
                updated.append(f"{label}: {info[field]}")
        if updated:
            self._log("✓ 从文件名自动提取: " + " | ".join(updated))

    def _extract(self):
        text = self.job_text.get("1.0", tk.END).strip()
        if not text:
            self._log("⚠ 请先粘贴招聘信息到窗口③")
            return
        info = extract_info(text)
        updates = []
        if info["email"]:
            self.to_email.set(info["email"])
            updates.append(f"邮箱: {info['email']}")
        if info["naming_format"]:
            raw_fmt = info["naming_format"]
            self.raw_naming_format = raw_fmt  # 保存原始格式，供 _send 使用
            filled_fmt = self._fill_template(raw_fmt, info)
            self.naming_format.set(filled_fmt)
            # 同时填充邮件标题（去掉.pdf等文件扩展名）
            subject = re.sub(r"\.pdf$", "", filled_fmt, flags=re.IGNORECASE).strip()
            self.email_subject.set(subject)
            if filled_fmt != raw_fmt:
                updates.append(f"命名格式: {raw_fmt} → {filled_fmt}")
            else:
                updates.append(f"命名格式: {filled_fmt}")
        if info["company"]:
            self.company_var.set(info["company"])
            updates.append(f"公司: {info['company']}")
        else:
            self.company_var.set("")
        if info["position"]:
            self.position_var.set(info["position"])
            updates.append(f"岗位: {info['position']}")
        else:
            self.position_var.set("")

        if updates:
            self._log("✓ 智能提取结果: " + " | ".join(updates))
            if not info["email"]:
                self._log("⚠ 未提取到邮箱，请手动填写")
            if not info["naming_format"]:
                self._log("⚠ 未提取到命名格式，请手动填写")
        else:
            self._log("⚠ 未能提取到有效信息，请手动填写各字段")

    def _send(self):
        self._update_sender_email()
        sender = self.sender_email.get().strip()
        auth = self.auth_code.get().strip()
        to = self.to_email.get().strip()
        fmt = self.naming_format.get().strip()
        resume = self.resume_path.get().strip()

        # 验证输入
        errors = []
        if not sender:
            errors.append("发件邮箱")
        if not auth:
            errors.append("授权码")
        if not to:
            errors.append("对方邮箱")
        if not fmt:
            errors.append("命名格式")
        if not resume:
            errors.append("原始简历")
        elif not os.path.exists(resume):
            errors.append("原始简历文件不存在")

        if errors:
            self._log(f"⚠ 请填写以下必填项: {', '.join(errors)}")
            return

        if "@" not in to:
            self._log("⚠ 对方邮箱格式不正确")
            return

        # 提取招聘信息用于模板填充
        text = self.job_text.get("1.0", tk.END).strip()
        info = extract_info(text)

        # 用真实信息填充命名模板中的占位符
        # 优先使用 _extract 保存的原始格式（未填充），避免二次填充导致乱码
        if self.raw_naming_format:
            raw_fmt = self.raw_naming_format
        else:
            raw_fmt = fmt.strip()
        filled_fmt = self._fill_template(raw_fmt, info)
        if filled_fmt != raw_fmt:
            self._log(f"✓ 模板填充: {raw_fmt} → {filled_fmt}")

        # 生成目标文件名
        ext = os.path.splitext(resume)[1]
        target_name = filled_fmt
        if not target_name.lower().endswith(ext.lower()):
            target_name += ext
        os.makedirs(RESUMES_DIR, exist_ok=True)
        target_path = os.path.join(RESUMES_DIR, target_name)

        # 复制简历
        try:
            shutil.copy2(resume, target_path)
            self._log(f"✓ 简历已保存: {target_name}")
        except Exception as e:
            self._log(f"✗ 简历复制失败: {e}")
            return

        # 组装邮件内容并发起后台发送（含自动重试逻辑）
        company = info["company"] or ""
        position = info["position"] or ""
        subject = self.email_subject.get().strip() or "实习申请"
        body = "尊敬的招聘负责人，\n\n您好！\n\n请查收附件中的个人简历，期待您的回复。\n\n此致\n敬礼"

        self.send_btn.config(state=tk.DISABLED)

        # 初始化重试状态，最多自动重试3次
        self._retry_state = {
            "sender": sender, "auth": auth, "to": to, "fmt": fmt,
            "subject": subject, "body": body, "target_path": target_path,
            "company": company, "position": position,
            "attempt": 0, "max_retries": 3,
        }

        self._log(f"正在发送邮件到 {to} ...")
        self._attempt_send()

    def _attempt_send(self):
        """在后台线程中尝试发送邮件。"""
        state = self._retry_state
        if state is None:
            return

        def do_send():
            try:
                send_email(
                    state["sender"], state["auth"], state["to"],
                    state["subject"], state["body"], state["target_path"],
                )
                self.root.after(0, self._on_send_success,
                    state["company"], state["position"], state["to"], state["fmt"])
            except Exception as e:
                self.root.after(0, self._on_attempt_failed, str(e))

        threading.Thread(target=do_send, daemon=True).start()

    def _on_attempt_failed(self, error_msg):
        """一次发送失败后的处理：还有重试次数则调度重试，否则最终失败。"""
        state = self._retry_state
        if state is None:
            return

        state["attempt"] += 1
        attempt = state["attempt"]
        max_retries = state["max_retries"]

        if attempt <= max_retries:
            self._log(f"⚠ 第{attempt}次发送失败: {error_msg}，3秒后自动重试...")
            self._schedule_retry()
        else:
            self._retry_state = None
            self._on_send_error(error_msg)

    def _schedule_retry(self):
        """显示取消按钮，3秒后自动重试。"""
        self.cancel_retry_btn.pack(side=tk.LEFT, padx=(10, 0))
        self._retry_after_id = self.root.after(3000, self._do_retry)

    def _do_retry(self):
        """执行重试：隐藏取消按钮，重新发送。"""
        self._retry_after_id = None
        self.cancel_retry_btn.pack_forget()
        self._attempt_send()

    def _cancel_retry(self):
        """用户点击取消重试按钮。"""
        if self._retry_after_id is not None:
            self.root.after_cancel(self._retry_after_id)
            self._retry_after_id = None
        self.cancel_retry_btn.pack_forget()
        state = self._retry_state
        self._retry_state = None
        if state:
            self._log("⚠ 用户取消重试")
            self._on_send_error("用户取消重试")

    def _on_send_success(self, company, position, to_email, naming_fmt):
        self._retry_state = None
        self._hide_retry_button()
        self.send_btn.config(state=tk.NORMAL)
        self._log(f"✓ 邮件发送成功!")

        # 保存历史
        record = {
            "company": company or "未知",
            "position": position or "未知",
            "email": to_email,
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "naming_format": naming_fmt,
            "resume_file": naming_fmt + ".pdf",
            "stages": {
                "已投递": "pass",
                "简历通过": "pending",
                "一面": "pending",
                "二面": "pending",
                "三面": "pending",
                "已Offer": "pending",
            },
        }
        self.records.insert(0, record)
        save_history(self.records)
        messagebox.showinfo("发送成功", f"邮件已发送至 {to_email}")

    def _hide_retry_button(self):
        if self._retry_after_id is not None:
            self.root.after_cancel(self._retry_after_id)
            self._retry_after_id = None
        self.cancel_retry_btn.pack_forget()

    def _on_send_error(self, error_msg):
        self._retry_state = None
        self._hide_retry_button()
        self.send_btn.config(state=tk.NORMAL)
        self._log(f"✗ 邮件发送失败（已重试3次仍失败）: {error_msg}")
        # 根据错误码给出针对性提示
        if "535" in error_msg or "authentication" in error_msg.lower():
            self._log("💡 提示: 请检查授权码是否正确。网易邮箱需在设置中生成授权码，非登录密码。")
        messagebox.showerror("发送失败", f"邮件发送失败（已自动重试3次）:\n{error_msg}")

    def _open_history(self):
        HistoryWindow(self.root, self.records)

    def _show_tutorial(self):
        win = tk.Toplevel(self.root)
        win.title("使用指南")
        win.geometry("480x420")
        win.resizable(True, True)
        win.configure(bg="#F0F0F0")
        win.transient(self.root)
        win.grab_set()

        canvas = tk.Canvas(win, bg="#F0F0F0", highlightthickness=0)
        scrollbar = tk.Scrollbar(win, orient=tk.VERTICAL, command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg="#F0F0F0")

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas_window = canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def _on_canvas_configure(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        main = tk.Frame(scroll_frame, bg="#F0F0F0")
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        tk.Label(main, text="使用指南", font=self.f_heading,
                 bg="#F0F0F0").pack(anchor=tk.W, pady=(0, 10))

        intro = (
            "这是一款帮你投递简历的软件。不同公司要求的简历命名格式不同，"
            "每次都要在文件夹里手动改名太麻烦了。\n"
            "这款软件可以一键自动命名并发送简历。"
        )
        tk.Label(main, text=intro, font=self.f_small,
                 fg="#555555", bg="#F0F0F0",
                 wraplength=400, justify=tk.LEFT).pack(anchor=tk.W, pady=(0, 14))

        steps = [
            ("配置发件信息", "填写你的邮箱和授权码。\n点击授权码旁的 ⓘ 查看获取方式"),
            ("填写个人信息", "软件会根据命名格式要求，自动填入\n姓名、学校、专业等信息"),
            ("选择原始简历", "在原始简历的基础上创建副本，\n原文件不会被修改"),
            ("粘贴招聘信息", "将包含公司、岗位、邮箱、命名格式的\n招聘JD全文粘贴进来"),
            ("智能提取", "自动识别命名格式、邮箱、公司、岗位"),
            ("一键发送", "核对无误后，自动重命名并发送邮件"),
        ]
        for i, (title, desc) in enumerate(steps):
            card = tk.LabelFrame(main, bg="#F0F0F0", padx=10, pady=8)
            card.pack(fill=tk.X, pady=(0, 6))
            num = tk.Label(card, text=str(i + 1),
                           font=tkfont.Font(family="Microsoft YaHei", size=12, weight="bold"),
                           fg="#0066CC", bg="#F0F0F0")
            num.pack(side=tk.LEFT, padx=(0, 10))
            tk.Label(card, text=title,
                     font=self.f_body,
                     bg="#F0F0F0").pack(anchor=tk.W)
            tk.Label(card, text=desc,
                     font=self.f_small,
                     fg="#555555", bg="#F0F0F0").pack(anchor=tk.W, pady=(2, 0))

        tk.Button(main, text="知道了", font=self.f_body, width=10,
                  command=win.destroy).pack(pady=(6, 0))

    def _on_close(self):
        # 窗口关闭前保存所有配置信息
        self._update_sender_email()
        self.config["email"] = self.sender_email.get().strip()
        self.config["auth_code"] = self.auth_code.get().strip()
        self.config["resume_path"] = self.resume_path.get().strip()
        self.config["personal"] = {
            "name": self.user_name.get().strip(),
            "phone": self.user_phone.get().strip(),
            "school": self.user_school.get().strip(),
            "internship_duration": self.internship_duration.get().strip(),
            "arrival_time": self.arrival_time.get().strip(),
            "weekly_attendance": self.weekly_attendance.get().strip(),
            "undergrad_school": self.undergrad_school.get().strip(),
            "grad_school": self.grad_school.get().strip(),
            "undergrad_major": self.undergrad_major.get().strip(),
            "grad_major": self.grad_major.get().strip(),
            "graduation_time": self.graduation_time.get().strip(),
            "grade": self.user_grade.get().strip(),
        }
        save_config(self.config)

        # 如果代码有改动，关闭时自动后台更新 exe
        self._auto_build_exe()

        self.root.destroy()

    def _auto_build_exe(self):
        # 仅在源码运行且代码有改动时自动更新 exe
        if getattr(sys, "frozen", False):
            return  # 已是 exe，无需自更新
        build_script = os.path.join(BASE_DIR, "build_share.py")
        main_py = os.path.join(BASE_DIR, "main.py")
        exe_path = os.path.join(os.path.dirname(BASE_DIR), "智能投递助手.exe")
        if not os.path.exists(build_script):
            return
        if os.path.exists(exe_path) and os.path.getmtime(main_py) <= os.path.getmtime(exe_path):
            return
        self._log("代码已改动，正在更新 exe 文件（约30秒），请稍等...")
        subprocess.Popen(
            [sys.executable, build_script],
            creationflags=subprocess.CREATE_NEW_CONSOLE,
        )


if __name__ == "__main__":
    app = App()
    app.root.mainloop()
